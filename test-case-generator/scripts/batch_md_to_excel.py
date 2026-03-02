"""
批量Markdown测试用例转Excel工具

功能：
1. 批量解析case_MD目录下的所有Markdown测试用例文件
2. 将所有测试用例合并到一个Excel文件中
3. 按模块分组，生成连续的用例编号

用法：
    python batch_md_to_excel.py [输出excel路径] [--prefix 前缀]

示例：
    python batch_md_to_excel.py
    python batch_md_to_excel.py case_excel/所有测试用例.xlsx
    python batch_md_to_excel.py --prefix TEST
"""

import argparse
import re
import sys
import os
from dataclasses import dataclass
from typing import List, Optional
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装openpyxl库")
    print("请运行：pip install openpyxl")
    sys.exit(1)


# 模块名称到缩写的映射
MODULE_ABBR_MAP = {
    # 不带"测试用例"后缀的模块名称映射
    "唯一码收货": "WYMSH",
    "唯一码组盘": "WYMZP",
    "唯一码直接盘点": "WYMZHPD",
    "唯一码按任务盘点": "WYMRWPD",
    "唯一码直接移库": "WYMZHYK",
    "唯一码直接补货": "WYMZHBH",
    "唯一码按任务拣货": "WYMRWLH",
    "唯一码在线拣选": "WYMZXJX",
    "回流入库": "HLRK",
    "唯一码自动分配": "WYMZDFP",
    "唯一码手动分配": "WYMSDFP",
    "唯一码快捷出库": "WYMKJCK",
    "唯一码库存调整": "WYMKCTZ",
    "生成上架任务": "SCSJRW",
    "通用测试用例": "TY",
    # 带"测试用例"后缀的模块名称映射
    "唯一码收货测试用例": "WYMSH",
    "唯一码组盘测试用例": "WYMZP",
    "唯一码直接盘点测试用例": "WYMZHPD",
    "唯一码按任务盘点测试用例": "WYMRWPD",
    "唯一码直接移库测试用例": "WYMZHYK",
    "唯一码直接补货测试用例": "WYMZHBH",
    "唯一码按任务拣货测试用例": "WYMRWLH",
    "唯一码在线拣选测试用例": "WYMZXJX",
    "回流入库测试用例": "HLRK",
    "唯一码自动分配测试用例": "WYMZDFP",
    "唯一码手动分配测试用例": "WYMSDFP",
    "唯一码快捷出库测试用例": "WYMKJCK",
    "唯一码库存调整测试用例": "WYMKCTZ",
    "生成上架任务测试用例": "SCSJRW",
}


def get_module_abbr(module_name: str) -> str:
    """
    获取模块名称的缩写
    优先从映射表查找，未找到则返回原名称的大写形式
    """
    return MODULE_ABBR_MAP.get(module_name, module_name.upper())


def format_content_with_numbering(content: str) -> str:
    """
    为内容添加序号编号
    单行内容不添加，多行内容每行添加序号
    """
    if not content:
        return content

    lines = content.split("\n")

    # 单行内容直接返回
    if len(lines) <= 1:
        return content

    # 多行内容添加序号
    numbered_lines = []
    for i, line in enumerate(lines, 1):
        stripped_line = line.strip()
        if stripped_line:  # 只处理非空行
            numbered_lines.append(f"{i}. {stripped_line}")
        else:
            numbered_lines.append(line)

    return "\n".join(numbered_lines)


@dataclass
class TestCase:
    """测试用例数据类"""

    module: str  # 所属模块
    feature: str  # 功能点
    priority: str  # 优先级 P0/P1/P2/P3
    title: str  # 用例标题（来自###标题行）
    precondition: str  # 前置条件
    steps: str  # 操作步骤
    expected: str  # 预期结果
    case_no: str = ""  # 用例编号
    source_file: str = ""  # 来源文件


class MarkdownParser:
    """Markdown测试用例解析器"""

    def __init__(self, content: str, source_file: str = ""):
        self.content = content
        self.lines = content.split("\n")
        self.current_module = ""
        self.current_feature = ""
        self.current_case_title = ""
        self.cases: List[TestCase] = []
        self.source_file = source_file

    def parse(self) -> List[TestCase]:
        """解析Markdown内容，提取测试用例"""
        i = 0
        while i < len(self.lines):
            line = self.lines[i].strip()
            if not line:
                i += 1
                continue

            # 解析模块 (# 开头，但排除##和###)
            if line.startswith("# ") and not line.startswith("## "):
                self.current_module = line[2:].strip()
                self.current_feature = ""
                self.current_case_title = ""

            # 解析功能点 (## 开头)
            elif line.startswith("## "):
                self.current_feature = line[3:].strip()
                self.current_case_title = ""

            # 解析测试用例标题 (### 开头)
            elif line.startswith("### "):
                self.current_case_title = line[4:].strip()

            # 解析测试用例内容 (- [Px] 开头)
            elif line.startswith("- [P") and self.current_case_title:
                case = self._parse_case_line(line)
                if case:
                    self.cases.append(case)

            i += 1

        return self.cases

    def _parse_case_line(self, line: str) -> Optional[TestCase]:
        """
        解析单行测试用例
        格式：- [Px] 前置条件 | 操作步骤 | 预期结果
        """
        # 提取优先级
        priority_match = re.match(r"- \[(P\d+)\] (.+)", line)
        if not priority_match:
            return None

        priority = priority_match.group(1)
        remaining = priority_match.group(2)

        # 按 | 分割字段
        parts = [p.strip() for p in remaining.split("|")]

        if len(parts) >= 3:
            precondition = parts[0].replace(";", "\n").replace("；", "\n")
            steps = parts[1].replace(";", "\n").replace("；", "\n")
            expected = parts[2].replace(";", "\n").replace("；", "\n")
        elif len(parts) == 2:
            precondition = parts[0].replace(";", "\n").replace("；", "\n")
            steps = parts[1].replace(";", "\n").replace("；", "\n")
            expected = ""
        elif len(parts) == 1:
            precondition = parts[0].replace(";", "\n").replace("；", "\n")
            steps = ""
            expected = ""
        else:
            precondition = ""
            steps = ""
            expected = ""

        return TestCase(
            module=self.current_module,
            feature=self.current_feature,
            priority=priority,
            title=self.current_case_title,
            precondition=precondition,
            steps=steps,
            expected=expected,
            source_file=self.source_file,
        )


def generate_case_numbers(cases: List[TestCase], prefix: str = "DH") -> None:
    """为所有用例生成编号"""
    module_counter = {}
    for case in cases:
        if case.module not in module_counter:
            module_counter[case.module] = 0
        module_counter[case.module] += 1

        module_abbr = get_module_abbr(case.module)
        case.case_no = f"{prefix}_{module_abbr}_{module_counter[case.module]:03d}"


class ExcelGenerator:
    """Excel测试用例生成器"""

    def __init__(self, cases: List[TestCase]):
        self.cases = cases
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "测试用例"

    def generate(self, output_path: str):
        """生成Excel文件"""
        self._setup_header()
        self._fill_data()
        self._apply_styles()
        self._adjust_column_widths()
        self.wb.save(output_path)
        print(f"✅ Excel文件已生成：{output_path}")
        print(f"📊 共 {len(self.cases)} 条测试用例")

    def _setup_header(self):
        """设置表头"""
        headers = [
            "用例编号",
            "所属模块",
            "功能点",
            "优先级",
            "用例标题",
            "前置条件",
            "操作步骤",
            "预期结果",
            "来源文件",
        ]

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    def _fill_data(self):
        """填充数据"""
        for row, case in enumerate(self.cases, 2):
            self.ws.cell(row=row, column=1, value=case.case_no)
            self.ws.cell(row=row, column=2, value=case.module)
            self.ws.cell(row=row, column=3, value=case.feature)
            self.ws.cell(row=row, column=4, value=case.priority)
            self.ws.cell(row=row, column=5, value=case.title)
            self.ws.cell(
                row=row,
                column=6,
                value=format_content_with_numbering(case.precondition),
            )
            self.ws.cell(
                row=row, column=7, value=format_content_with_numbering(case.steps)
            )
            self.ws.cell(
                row=row, column=8, value=format_content_with_numbering(case.expected)
            )
            self.ws.cell(row=row, column=9, value=case.source_file)

    def _apply_styles(self):
        """应用样式"""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        priority_colors = {
            "P0": "FF6B6B",
            "P1": "FFD93D",
            "P2": "6BCB77",
            "P3": "95A5A6",
        }

        for row in range(2, len(self.cases) + 2):
            priority_cell = self.ws.cell(row=row, column=4)
            priority = priority_cell.value

            if priority in priority_colors:
                priority_cell.fill = PatternFill(
                    start_color=priority_colors[priority],
                    end_color=priority_colors[priority],
                    fill_type="solid",
                )
                priority_cell.font = Font(bold=True)

            for col in range(1, 10):
                cell = self.ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _adjust_column_widths(self):
        """调整列宽"""
        column_widths = {
            1: 18,  # 用例编号
            2: 20,  # 所属模块
            3: 25,  # 功能点
            4: 10,  # 优先级
            5: 40,  # 用例标题
            6: 35,  # 前置条件
            7: 40,  # 操作步骤
            8: 40,  # 预期结果
            9: 25,  # 来源文件
        }

        for col, width in column_widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = width

        self.ws.row_dimensions[1].height = 30
        for row in range(2, len(self.cases) + 2):
            self.ws.row_dimensions[row].height = 60


def get_all_md_files(case_dir: str) -> List[str]:
    """获取case_MD目录下的所有md文件，按文件名排序"""
    md_files = []
    for f in os.listdir(case_dir):
        if f.endswith(".md"):
            md_files.append(os.path.join(case_dir, f))
    # 按文件名排序
    md_files.sort(key=lambda x: os.path.basename(x))
    return md_files


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="批量Markdown测试用例转Excel工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  python batch_md_to_excel.py
  python batch_md_to_excel.py case_excel/所有测试用例.xlsx
  python batch_md_to_excel.py --prefix TEST
        """,
    )
    parser.add_argument("output", nargs="?", help="输出的Excel文件路径（可选）")
    parser.add_argument("--prefix", "-p", default="DH", help="用例编号前缀，默认为DH")
    parser.add_argument(
        "--case-dir", "-d", default="case_MD", help="测试用例目录，默认为case_MD"
    )

    args = parser.parse_args()

    prefix = args.prefix
    case_dir = args.case_dir

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"case_excel/所有测试用例_{timestamp}.xlsx"

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    # 获取所有md文件
    if not os.path.exists(case_dir):
        print(f"❌ 错误：目录不存在 - {case_dir}")
        sys.exit(1)

    md_files = get_all_md_files(case_dir)
    if not md_files:
        print(f"⚠️ 警告：在 {case_dir} 目录下未找到Markdown文件")
        sys.exit(1)

    print(f"📁 找到 {len(md_files)} 个Markdown文件：")
    for f in md_files:
        print(f"   - {os.path.basename(f)}")
    print()

    # 解析所有测试用例
    all_cases: List[TestCase] = []
    for md_path in md_files:
        print(f"📖 正在解析：{os.path.basename(md_path)}")
        try:
            with open(md_path, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            print(f"   ⚠️ 读取文件失败：{e}")
            continue

        parser = MarkdownParser(content, source_file=os.path.basename(md_path))
        cases = parser.parse()
        all_cases.extend(cases)
        print(f"   ✅ 找到 {len(cases)} 条测试用例")

    if not all_cases:
        print("\n❌ 错误：未找到任何测试用例")
        sys.exit(1)

    print(f"\n📊 总计：{len(all_cases)} 条测试用例")

    # 统计优先级
    priority_count = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    for case in all_cases:
        if case.priority in priority_count:
            priority_count[case.priority] += 1

    print(f"   P0（高）：{priority_count['P0']} 条")
    print(f"   P1（中）：{priority_count['P1']} 条")
    print(f"   P2（低）：{priority_count['P2']} 条")
    print(f"   P3（共享）：{priority_count['P3']} 条")

    # 生成用例编号
    generate_case_numbers(all_cases, prefix)

    # 生成Excel
    print(f"\n📝 正在生成Excel：{output_path}")
    generator = ExcelGenerator(all_cases)
    generator.generate(output_path)

    print("\n✨ 转换完成！")


if __name__ == "__main__":
    main()
