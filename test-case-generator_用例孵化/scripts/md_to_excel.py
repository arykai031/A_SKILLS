"""
Markdown测试用例转Excel工具

功能：
1. 解析Markdown格式的测试用例
2. 提取模块、功能点、用例信息
3. 生成结构化的Excel测试用例文档

用法：
    python md_to_excel.py <markdown文件路径> [输出excel路径]

示例：
    python md_to_excel.py CaseMD/唯一码收货_测试用例.md
    python md_to_excel.py CaseMD/唯一码收货测试用例.md CaseExcel/唯一码收货_测试用例.xlsx
"""

import argparse
import re
import sys
import os
from dataclasses import dataclass
from typing import List, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装openpyxl库")
    print("请运行：pip install openpyxl")
    sys.exit(1)


# 模块名称到缩写的映射
MODULE_ABBR_MAP = {
    "唯一码收货": "WYMSH",
    "唯一码组盘": "WYMZP",
    "唯一码直接盘点": "WYMZHPD",
    "唯一码按任务盘点": "WYMRWPD",
    "唯一码直接移库": "WYMZHYK",
    "唯一码直接补货": "WYMZHBH",
    "唯一码按任务拣货": "WYMRWLH",
    "唯一码在线拣选": "WYMZXJX",
    "回流入库": "HLRK",
    "唯一码自动分配": "WYMZDFP",
    "唯一码手动分配": "WYMSDFP",
    "唯一码快捷出库": "WYMKJCK",
    "唯一码库存调整": "WYMKCTZ",
    "生成上架任务": "SCSJRW",
}


def get_module_abbr(module_name: str) -> str:
    """
    获取模块名称的缩写
    优先从映射表查找，未找到则返回原名称的大写形式
    """
    return MODULE_ABBR_MAP.get(module_name, module_name.upper())


def format_content_with_numbering(content: str) -> str:
    """
    为内容添加序号编号
    单行内容不添加，多行内容每行添加序号

    Args:
        content: 原始内容文本

    Returns:
        添加序号后的文本
    """
    if not content:
        return content

    lines = content.split("\n")

    # 单行内容直接返回
    if len(lines) <= 1:
        return content

    # 多行内容添加序号
    numbered_lines = []
    for i, line in enumerate(lines, 1):
        stripped_line = line.strip()
        if stripped_line:  # 只处理非空行
            numbered_lines.append(f"{i}. {stripped_line}")
        else:
            numbered_lines.append(line)

    return "\n".join(numbered_lines)


@dataclass
class TestCase:
    """测试用例数据类"""

    module: str  # 所属模块
    feature: str  # 功能点
    priority: str  # 优先级 P0/P1/P2/P3
    title: str  # 用例标题（来自###标题行）
    precondition: str  # 前置条件
    steps: str  # 操作步骤
    expected: str  # 预期结果
    case_no: str = ""  # 用例编号


class MarkdownParser:
    """Markdown测试用例解析器"""

    def __init__(self, content: str):
        self.content = content
        self.lines = content.split("\n")
        self.current_module = ""
        self.current_feature = ""
        self.current_case_title = ""  # 当前用例标题（来自###行）
        self.cases: List[TestCase] = []

    def parse(self, prefix: str = "DH") -> List[TestCase]:
        """解析Markdown内容，提取测试用例"""
        i = 0
        while i < len(self.lines):
            line = self.lines[i].strip()
            if not line:
                i += 1
                continue

            # 解析模块 (# 开头，但排除##和###)
            if line.startswith("# ") and not line.startswith("## "):
                self.current_module = line[2:].strip()
                self.current_feature = ""  # 重置功能点
                self.current_case_title = ""  # 重置用例标题

            # 解析功能点 (## 开头)
            elif line.startswith("## "):
                self.current_feature = line[3:].strip()
                self.current_case_title = ""  # 重置用例标题

            # 解析测试用例标题 (### 开头)
            elif line.startswith("### "):
                self.current_case_title = line[4:].strip()

            # 解析测试用例内容 (- [Px] 开头)
            elif line.startswith("- [P") and self.current_case_title:
                case = self._parse_case_line(line)
                if case:
                    self.cases.append(case)

            i += 1

        # 生成用例编号，传入前缀
        self._generate_case_numbers(prefix)
        return self.cases

    def _parse_case_line(self, line: str) -> Optional[TestCase]:
        """
        解析单行测试用例
        格式：- [Px] 前置条件 | 操作步骤 | 预期结果
        （用例标题来自###行）
        """
        # 提取优先级
        priority_match = re.match(r"- \[(P\d+)\] (.+)", line)
        if not priority_match:
            return None

        priority = priority_match.group(1)
        remaining = priority_match.group(2)

        # 按 | 分割字段（3个字段：前置条件、步骤、预期结果）
        parts = [p.strip() for p in remaining.split("|")]

        if len(parts) >= 3:
            # 格式：[Px] 前置条件 | 操作步骤 | 预期结果
            precondition = parts[0].replace(";", "\n")
            steps = parts[1].replace(";", "\n")
            expected = parts[2].replace(";", "\n")
        elif len(parts) == 2:
            # 只有前置条件和步骤
            precondition = parts[0].replace(";", "\n")
            steps = parts[1].replace(";", "\n")
            expected = ""
        elif len(parts) == 1:
            # 只有前置条件
            precondition = parts[0].replace(";", "\n")
            steps = ""
            expected = ""
        else:
            precondition = ""
            steps = ""
            expected = ""

        return TestCase(
            module=self.current_module,
            feature=self.current_feature,
            priority=priority,
            title=self.current_case_title,  # 使用###行的标题
            precondition=precondition,
            steps=steps,
            expected=expected,
        )

    def _generate_case_numbers(self, prefix: str = "DH"):
        """生成用例编号"""
        module_counter = {}
        for case in self.cases:
            if case.module not in module_counter:
                module_counter[case.module] = 0
            module_counter[case.module] += 1

            # 获取模块缩写
            module_abbr = get_module_abbr(case.module)

            # 生成编号: xxx_模块_序号
            case.case_no = f"{prefix}_{module_abbr}_{module_counter[case.module]:03d}"


class ExcelGenerator:
    """Excel测试用例生成器"""

    def __init__(self, cases: List[TestCase]):
        self.cases = cases
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "测试用例"

    def generate(self, output_path: str):
        """生成Excel文件"""
        self._setup_header()
        self._fill_data()
        self._apply_styles()
        self._adjust_column_widths()
        self.wb.save(output_path)
        print(f"✅ Excel文件已生成：{output_path}")
        print(f"📊 共 {len(self.cases)} 条测试用例")

    def _setup_header(self):
        """设置表头"""
        headers = [
            "用例编号",
            "所属模块",
            "功能点",
            "优先级",
            "用例标题",
            "前置条件",
            "操作步骤",
            "预期结果",
        ]

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    def _fill_data(self):
        """填充数据"""
        for row, case in enumerate(self.cases, 2):
            self.ws.cell(row=row, column=1, value=case.case_no)
            self.ws.cell(row=row, column=2, value=case.module)
            self.ws.cell(row=row, column=3, value=case.feature)
            self.ws.cell(row=row, column=4, value=case.priority)
            self.ws.cell(row=row, column=5, value=case.title)
            # 对前置条件、操作步骤、预期结果添加序号格式化
            self.ws.cell(
                row=row,
                column=6,
                value=format_content_with_numbering(case.precondition),
            )
            self.ws.cell(
                row=row, column=7, value=format_content_with_numbering(case.steps)
            )
            self.ws.cell(
                row=row, column=8, value=format_content_with_numbering(case.expected)
            )

    def _apply_styles(self):
        """应用样式"""
        # 定义边框
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 优先级颜色映射
        priority_colors = {
            "P0": "FF6B6B",  # 红色 - 高优先级
            "P1": "FFD93D",  # 黄色 - 中优先级
            "P2": "6BCB77",  # 绿色 - 低优先级
            "P3": "95A5A6",  # 灰色 - 最低优先级（共享用例）
        }

        for row in range(2, len(self.cases) + 2):
            # 获取优先级单元格
            priority_cell = self.ws.cell(row=row, column=4)
            priority = priority_cell.value

            # 根据优先级设置背景色
            if priority in priority_colors:
                priority_cell.fill = PatternFill(
                    start_color=priority_colors[priority],
                    end_color=priority_colors[priority],
                    fill_type="solid",
                )
                priority_cell.font = Font(bold=True)

            # 应用边框和对齐
            for col in range(1, 9):
                cell = self.ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _adjust_column_widths(self):
        """调整列宽"""
        column_widths = {
            1: 15,  # 用例编号
            2: 20,  # 所属模块
            3: 25,  # 功能点
            4: 10,  # 优先级
            5: 40,  # 用例标题
            6: 35,  # 前置条件
            7: 40,  # 操作步骤
            8: 40,  # 预期结果
        }

        for col, width in column_widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = width

        # 设置行高
        self.ws.row_dimensions[1].height = 30
        for row in range(2, len(self.cases) + 2):
            self.ws.row_dimensions[row].height = 60


def ensure_directory(path: str) -> str:
    """确保目录存在，如果不存在则创建"""
    directory = os.path.dirname(path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
    return path


def get_default_output_path(input_path: str) -> str:
    """根据输入路径生成默认输出路径"""
    input_path_obj = Path(input_path)

    # 如果输入路径包含 CaseMD，则输出到 CaseExcel
    if "CaseMD" in input_path_obj.parts:
        # 将 CaseMD 替换为 CaseExcel
        parts = list(input_path_obj.parts)
        for i, part in enumerate(parts):
            if part == "CaseMD":
                parts[i] = "CaseExcel"
                break
        output_path = Path(*parts).with_suffix(".xlsx")
        return str(output_path)

    # 默认在同一目录下，扩展名改为.xlsx
    return str(input_path_obj.with_suffix(".xlsx"))


def main():
    """主函数"""
    # 解析命令行参数
    parser = argparse.ArgumentParser(
        description="Markdown测试用例转Excel工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  python md_to_excel.py CaseMD/唯一码收货测试用例.md
  python md_to_excel.py CaseMD/唯一码收货测试用例.md CaseExcel/唯一码收货测试用例.xlsx
  python md_to_excel.py CaseMD/唯一码收货测试用例.md --prefix TEST
        """,
    )
    parser.add_argument("input", help="输入的Markdown文件路径")
    parser.add_argument("output", nargs="?", help="输出的Excel文件路径（可选）")
    parser.add_argument("--prefix", "-p", default="DH", help="用例编号前缀，默认为DH")

    args = parser.parse_args()

    md_path = args.input
    prefix = args.prefix

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        output_path = get_default_output_path(md_path)

    # 检查输入文件是否存在
    if not os.path.exists(md_path):
        print(f"❌ 错误：文件不存在 - {md_path}")
        sys.exit(1)

    # 确保输出目录存在
    ensure_directory(output_path)

    # 读取Markdown文件
    try:
        with open(md_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        print(f"❌ 错误：读取文件失败 - {e}")
        sys.exit(1)

    # 解析测试用例，传入前缀
    print(f"📖 正在解析：{md_path}")
    parser = MarkdownParser(content)
    cases = parser.parse(prefix=prefix)

    if not cases:
        print("⚠️ 警告：未找到测试用例，请检查文件格式")
        sys.exit(1)

    print(f"✅ 找到 {len(cases)} 条测试用例")

    # 统计优先级
    priority_count = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    for case in cases:
        if case.priority in priority_count:
            priority_count[case.priority] += 1

    print(f"   P0（高）：{priority_count['P0']} 条")
    print(f"   P1（中）：{priority_count['P1']} 条")
    print(f"   P2（低）：{priority_count['P2']} 条")
    print(f"   P3（共享）：{priority_count['P3']} 条")

    # 生成Excel
    print(f"\n📝 正在生成Excel：{output_path}")
    generator = ExcelGenerator(cases)
    generator.generate(output_path)


if __name__ == "__main__":
    main()
